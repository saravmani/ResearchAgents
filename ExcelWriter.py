"""
ExcelWriter.py

This script reads an Excel file and a JSON configuration, identifies structured metric rows based on header rules,
matches them with JSON metric names, and updates specified cells in-place.

Features:
- Parses structured Excel rows using configurable delimiters.
- Supports metric name reconstruction from complex header + value.
- Logs matched/unmatched values.
- Prints a final summary of matched/unmatched metrics.
- Includes robust error handling and logging for production use.
"""
import pandas as pd
import numpy as np
import string
import json
from openpyxl import load_workbook
import logging
logging.basicConfig(level=logging.INFO, format="%(message)s")

def clean_string(s, cleanup_chars=[" ", "-", "_"]):
    for char in cleanup_chars:
        s = s.replace(char, "")
    return s.lower()

def process_excel_with_json(
    excel_path, json_path, delimiter="<<"
):
    """
    Process an Excel file and update specific cells based on JSON metric mappings.

    Args:
        excel_path (str): Path to the Excel file.
        json_path (str): Path to the JSON file containing metricName and metricValue.
        delimiter (str): Delimiter used to combine header and value (e.g., "<<").

    Returns:
        None
    """
    try:
        try:
            with open(json_path, "r") as f:
                json_data = json.load(f)
        except Exception as e:
            logging.exception(f"❌ Failed to load or parse JSON file: {e}")
            return

        metric_col_letter = json_data["metric_column_name"]
        update_col_letter = json_data["column_To_Update_column_name"]
        metrics = json_data["data"]

        # Convert letters to column index (0-based)
        metric_col_index = string.ascii_uppercase.index(metric_col_letter.upper())
        update_col_index = string.ascii_uppercase.index(update_col_letter.upper())
        start_col_index = metric_col_index  # use metric_column_name for header and data lookup

        # Load Excel
        wb = load_workbook(excel_path)
        ws = wb.active

        # Read and trim cell values
        data = []
        for row in ws.iter_rows(values_only=True):
            trimmed = [str(cell).strip() if cell else "" for cell in row]
            data.append(trimmed)
        df = pd.DataFrame(data)

        current_header = None
        group_value = None
        combined_records = []

        for i in range(len(df)):
            row = df.iloc[i]
            after_b = row[start_col_index:]

            # Detect header row: only first column filled (ignore rest or zeros)
            if (
                row[start_col_index] and
                all(str(cell).strip() in ["", "0", "0.0"] for cell in row[start_col_index+1:])
            ):
                # Check if previous row is empty → stop header
                if i > 0:
                    prev_row = df.iloc[i-1]
                    if all(str(cell).strip() == "" for cell in prev_row) and all(str(cell).strip() == "" for cell in row):
                        current_header = None
                        continue
                current_header = row[start_col_index]
                continue

            # Updated logic for handling previous empty row and standalone metric names
            if i > 0:
                prev_row = df.iloc[i-1]
                if all(str(cell).strip() == "" for cell in prev_row):
                    current_header = None

            # Regardless of previous row, if no active header exists but value exists, treat it as standalone
            if not current_header and row[start_col_index]:
                value_b = row[start_col_index]
                combined_key = value_b
                combined_records.append({
                    "row_index": i + 1,
                    "header": None,
                    "value": combined_key
                })
                continue

            # Build combined key: Header<<Value in B
            if current_header and row[start_col_index]:
                value_b = row[start_col_index]
                combined_key = f"{current_header}{delimiter}{value_b}"
                combined_records.append({
                    "row_index": i + 1,
                    "header": current_header,
                    "value": combined_key
                })

        # Map metricName → metricValue from JSON
        metric_list = [{"key": d["metricName"], "value": d["metricValue"], "used": False} for d in metrics]

        logging.info("📋 Parsed Records from Excel:")
        for record in combined_records:
            logging.info(f"Row {record['row_index']}: {record['value']}")

        matched_json_keys = set()
        for record in combined_records:
            val = record["value"]
            val_clean = clean_string(val)
            matched = False
            for item in metric_list:
                json_key = item["key"]
                json_value = item["value"]
                if item["used"]:
                    continue
                if delimiter in json_key:
                    split_parts = json_key.split(delimiter)
                    if len(split_parts) >= 2:
                        reconstructed = f"{split_parts[-2]}{delimiter}{split_parts[-1]}"
                        json_key_clean = clean_string(json_key)
                        if val_clean in json_key_clean:
                            row_num = record["row_index"]
                            ws.cell(row=row_num, column=update_col_index + 1).value = json_value
                            logging.info(f"✅ Match: Row {row_num} updated with '{json_value}' for key '{val}'")
                            matched_json_keys.add(json_key)
                            matched = True
                            item["used"] = True
                            break
            if not matched:
                logging.info(f"❌ No match found for row {record['row_index']} with key '{val}'")


        # JSON Summary: matched and unmatched keys
        json_matched = []
        json_unmatched = []

        for item in metric_list:
            json_key = item["key"]
            json_key_clean = clean_string(json_key)
            is_matched = False
            for record in combined_records:
                val = record["value"]
                val_clean = clean_string(val)
                if val_clean in json_key_clean:
                    is_matched = True
                    break
            if is_matched:
                json_matched.append(json_key)
            else:
                json_unmatched.append(json_key)

        logging.info("📊 JSON Summary")
        logging.info(f"✅ Matched metrics in JSON: {len(json_matched)}")
        for key in json_matched:
            logging.info(f"  - {key}")
        logging.info(f"❌ Unmatched metrics in JSON: {len(json_unmatched)}")
        for key in json_unmatched:
            logging.info(f"  - {key}")

        logging.info("🧾 Audit Log of Matches:")
        for record in combined_records:
            val = record["value"]
            val_clean = clean_string(val)
            row_num = record["row_index"]
            for item in metric_list:
                json_key = item["key"]
                json_value = item["value"]
                json_key_clean = clean_string(json_key)
                if val_clean in json_key_clean:
                    logging.info(f"Row {row_num}: Excel key '{val}' matched with JSON key '{json_key}' → updated with '{json_value}'")
                    break


        # Save output
        # out_path = excel_path.replace(".xlsx", "_updated.xlsx")
        wb.save(excel_path)
        logging.info(f"✅ Updated file saved to: {excel_path} using delimiter '{delimiter}'")
    except Exception as e:
        logging.exception(f"❌ An error occurred while processing: {e}")


# Example usage
if __name__ == "__main__":
    process_excel_with_json(
        excel_path="input.xlsx",
        json_path="input.json",
        delimiter="<<"
    )


# Wrapper to process Excel with direct JSON string input and column names.
def process_excel_with_json_inputs(excel_path, json_string, metric_column_name, column_to_update_column_name, delimiter):
    """
    Wrapper to process Excel using direct JSON string input and column names.

    Args:
        excel_path (str): Path to the Excel file.
        json_string (str): JSON string containing metricName and metricValue list.
        metric_column_name (str): Column letter for metrics (e.g., 'B').
        column_to_update_column_name (str): Column letter where updates are to be made (e.g., 'D').
        delimiter (str): Delimiter used to combine header and value (e.g., '<<').

    Returns:
        None
    """
    json_data = json.loads(json_string)
    json_data["metric_column_name"] = metric_column_name
    json_data["column_To_Update_column_name"] = column_to_update_column_name
    return process_excel_with_json(excel_path, json_data, delimiter)