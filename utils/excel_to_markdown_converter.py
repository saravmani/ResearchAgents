import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from tabulate import tabulate
import re
from typing import List, Dict, Tuple, Optional


class ExcelToMarkdownConverter:
    """
    A comprehensive utility to convert Excel tables to Markdown format.
    Handles multiple tables, complex formatting, merged cells, and various table structures.
    """
    
    def __init__(self, excel_file_path: str):
        """
        Initialize the converter with an Excel file.
        
        Args:
            excel_file_path (str): Path to the Excel file
        """
        self.excel_file_path = excel_file_path
        self.workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    
    def convert_sheet_to_markdown(self, sheet_name: Optional[str] = None) -> str:
        """
        Convert an entire Excel sheet to Markdown format, detecting multiple tables.
        
        Args:
            sheet_name (str, optional): Name of the sheet to convert. If None, uses active sheet.
            
        Returns:
            str: Markdown formatted content with all detected tables
        """
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {self.workbook.sheetnames}")
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.active
        
        tables = self._detect_and_extract_tables(sheet)
        markdown_output = []
        
        if not tables:
            return "No tables detected in the Excel sheet."
        
        for i, table in enumerate(tables, 1):
            if table['markdown'].strip():
                markdown_output.append(f"## Table {i}")
                if table['title']:
                    markdown_output.append(f"**Title:** {table['title']}")
                markdown_output.append(f"**Range:** {table['range']}")
                markdown_output.append("")
                markdown_output.append(table['markdown'])
                markdown_output.append("")
        
        return "\n".join(markdown_output)
    
    def _detect_and_extract_tables(self, sheet) -> List[Dict]:
        """
        Detect and extract multiple tables from an Excel sheet.
        
        Args:
            sheet: openpyxl worksheet object
            
        Returns:
            List[Dict]: List of table dictionaries with data and markdown
        """
        tables = []
        processed_cells = set()
        
        # First pass: find all non-empty regions
        data_regions = self._find_data_regions(sheet)
        
        for region in data_regions:
            # Skip if this region overlaps with already processed cells
            if self._region_overlaps_processed(region, processed_cells):
                continue
            
            # Extract table data from this region
            table_data = self._extract_table_from_region(sheet, region)
            
            if table_data['data'] and len(table_data['data']) > 0:
                # Convert to markdown
                markdown_table = self._convert_data_to_markdown(table_data['data'])
                
                tables.append({
                    'range': f"{get_column_letter(region['min_col'])}{region['min_row']}:{get_column_letter(region['max_col'])}{region['max_row']}",
                    'title': table_data.get('title', ''),
                    'data': table_data['data'],
                    'markdown': markdown_table
                })
                
                # Mark cells as processed
                for row in range(region['min_row'], region['max_row'] + 1):
                    for col in range(region['min_col'], region['max_col'] + 1):
                        processed_cells.add((row, col))
        
        return tables
    
    def _find_data_regions(self, sheet) -> List[Dict]:
        """
        Find all continuous data regions in the sheet.
        
        Args:
            sheet: openpyxl worksheet object
            
        Returns:
            List[Dict]: List of region boundaries
        """
        regions = []
        visited = set()
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                if (row, col) not in visited:
                    cell = sheet.cell(row, col)
                    if cell.value is not None and str(cell.value).strip():
                        # Found start of a new region
                        region = self._expand_region(sheet, row, col, visited)
                        if region and self._is_valid_table_region(sheet, region):
                            regions.append(region)
        
        return regions
    
    def _expand_region(self, sheet, start_row: int, start_col: int, visited: set) -> Dict:
        """
        Expand a region from a starting cell to find connected data.
        
        Args:
            sheet: openpyxl worksheet object
            start_row (int): Starting row
            start_col (int): Starting column
            visited (set): Set of already visited cells
            
        Returns:
            Dict: Region boundaries
        """
        min_row = max_row = start_row
        min_col = max_col = start_col
        
        # Find the extent of continuous data
        # Expand right
        col = start_col
        while col <= sheet.max_column:
            cell = sheet.cell(start_row, col)
            if cell.value is None or not str(cell.value).strip():
                break
            max_col = col
            col += 1
        
        # Expand down
        row = start_row
        while row <= sheet.max_row:
            # Check if this row has data in any of the columns we've identified
            has_data = False
            for c in range(min_col, max_col + 1):
                cell = sheet.cell(row, c)
                if cell.value is not None and str(cell.value).strip():
                    has_data = True
                    break
            
            if not has_data:
                # Check next few rows in case of gaps
                gap_found = True
                for gap_row in range(row + 1, min(row + 3, sheet.max_row + 1)):
                    for c in range(min_col, max_col + 1):
                        cell = sheet.cell(gap_row, c)
                        if cell.value is not None and str(cell.value).strip():
                            gap_found = False
                            break
                    if not gap_found:
                        break
                
                if gap_found:
                    break
            
            max_row = row
            row += 1
        
        # Mark cells as visited
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                visited.add((r, c))
        
        return {
            'min_row': min_row,
            'max_row': max_row,
            'min_col': min_col,
            'max_col': max_col
        }
    
    def _is_valid_table_region(self, sheet, region: Dict) -> bool:
        """
        Check if a region represents a valid table (has enough data).
        
        Args:
            sheet: openpyxl worksheet object
            region (Dict): Region boundaries
            
        Returns:
            bool: True if valid table region
        """
        # Must have at least 2 rows and 2 columns, or be a single-column list
        min_rows = 2
        min_cols = 1
        
        row_count = region['max_row'] - region['min_row'] + 1
        col_count = region['max_col'] - region['min_col'] + 1
        
        if row_count < min_rows or col_count < min_cols:
            return False
        
        # Check data density - at least 30% of cells should have data
        total_cells = row_count * col_count
        filled_cells = 0
        
        for row in range(region['min_row'], region['max_row'] + 1):
            for col in range(region['min_col'], region['max_col'] + 1):
                cell = sheet.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    filled_cells += 1
        
        return filled_cells / total_cells >= 0.3
    
    def _region_overlaps_processed(self, region: Dict, processed_cells: set) -> bool:
        """
        Check if a region overlaps with already processed cells.
        
        Args:
            region (Dict): Region boundaries
            processed_cells (set): Set of processed cell coordinates
            
        Returns:
            bool: True if there's overlap
        """
        for row in range(region['min_row'], region['max_row'] + 1):
            for col in range(region['min_col'], region['max_col'] + 1):
                if (row, col) in processed_cells:
                    return True
        return False
    
    def _extract_table_from_region(self, sheet, region: Dict) -> Dict:
        """
        Extract table data from a specific region.
        
        Args:
            sheet: openpyxl worksheet object
            region (Dict): Region boundaries
            
        Returns:
            Dict: Table data and metadata
        """
        data = []
        title = ""
        
        # Look for a title (merged cell or single cell above the table)
        title_row = region['min_row'] - 1
        if title_row >= 1:
            for col in range(region['min_col'], region['max_col'] + 1):
                cell = sheet.cell(title_row, col)
                if cell.value is not None and str(cell.value).strip():
                    title = str(cell.value).strip()
                    break
        
        # Extract data from the region
        for row in range(region['min_row'], region['max_row'] + 1):
            row_data = []
            has_data = False
            
            for col in range(region['min_col'], region['max_col'] + 1):
                cell = sheet.cell(row, col)
                value = cell.value
                
                # Handle different data types
                if value is not None:
                    if isinstance(value, (int, float)):
                        cell_text = str(value)
                    else:
                        cell_text = str(value).strip()
                    has_data = True
                else:
                    cell_text = ""
                
                # Check for formatting (bold, underline) to identify headers
                if cell.font and (cell.font.bold or cell.font.underline):
                    cell_text = f"**{cell_text}**" if cell_text else ""
                
                row_data.append(cell_text)
            
            # Only add rows with data
            if has_data or any(row_data):
                data.append(row_data)
        
        # Handle merged cells
        data = self._handle_merged_cells(sheet, region)
        
        return {
            'data': data,
            'title': title
        }
    
    def _convert_data_to_markdown(self, data: List[List[str]]) -> str:
        """
        Convert table data to Markdown format.
        
        Args:
            data (List[List[str]]): Table data as list of rows
            
        Returns:
            str: Markdown formatted table
        """
        if not data or len(data) == 0:
            return ""
        
        # Clean data - ensure all rows have the same number of columns
        max_cols = max(len(row) for row in data) if data else 0
        clean_data = []
        
        for row in data:
            # Pad row to max columns
            padded_row = row + [""] * (max_cols - len(row))
            clean_data.append(padded_row)
        
        # Identify header rows (rows with bold text or first row)
        header_rows = []
        data_rows = []
        
        for i, row in enumerate(clean_data):
            is_header = False
            if i == 0:  # First row is usually header
                is_header = True
            else:
                # Check if row has bold formatting
                bold_count = sum(1 for cell in row if cell.startswith("**") and cell.endswith("**"))
                if bold_count > len(row) * 0.5:  # More than half cells are bold
                    is_header = True
            
            if is_header:
                # Clean bold markdown from headers
                clean_row = [cell.replace("**", "") for cell in row]
                header_rows.append(clean_row)
            else:
                data_rows.append(row)
        
        # Build markdown table
        markdown_lines = []
        
        # Use the last header row as column headers
        if header_rows:
            headers = header_rows[-1]
            markdown_lines.append("| " + " | ".join(headers) + " |")
            markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        else:
            # No headers found, use generic headers
            headers = [f"Column {i+1}" for i in range(max_cols)]
            markdown_lines.append("| " + " | ".join(headers) + " |")
            markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            # Treat all data as data rows
            data_rows = clean_data
        
        # Add data rows
        for row in data_rows:
            # Clean any remaining bold formatting from data
            clean_row = [cell.replace("**", "") for cell in row]
            markdown_lines.append("| " + " | ".join(clean_row) + " |")
        
        return "\n".join(markdown_lines)
    
    def convert_specific_range(self, range_str: str, sheet_name: Optional[str] = None) -> str:
        """
        Convert a specific Excel range to Markdown.
        
        Args:
            range_str (str): Excel range like "A1:D10"
            sheet_name (str, optional): Sheet name
            
        Returns:
            str: Markdown formatted table
        """
        if sheet_name:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.active
        
        # Parse range
        if ":" not in range_str:
            raise ValueError("Range must be in format 'A1:D10'")
        
        start_cell, end_cell = range_str.split(":")
        
        # Convert to row/column numbers
        start_col = column_index_from_string(re.match(r"[A-Z]+", start_cell).group())
        start_row = int(re.search(r"\d+", start_cell).group())
        end_col = column_index_from_string(re.match(r"[A-Z]+", end_cell).group())
        end_row = int(re.search(r"\d+", end_cell).group())
        
        region = {
            'min_row': start_row,
            'max_row': end_row,
            'min_col': start_col,
            'max_col': end_col
        }
        
        table_data = self._extract_table_from_region(sheet, region)
        return self._convert_data_to_markdown(table_data['data'])
    
    def get_sheet_names(self) -> List[str]:
        """
        Get all sheet names in the Excel file.
        
        Returns:
            List[str]: List of sheet names
        """
        return self.workbook.sheetnames
    
    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
    
    def _handle_merged_cells(self, sheet, region: Dict) -> List[List[str]]:
        """
        Handle merged cells in Excel and create appropriate table structure.
        
        Args:
            sheet: openpyxl worksheet object
            region (Dict): Region boundaries
            
        Returns:
            List[List[str]]: Processed table data with merged cells handled
        """
        data = []
        merged_ranges = list(sheet.merged_cells.ranges)
        
        for row in range(region['min_row'], region['max_row'] + 1):
            row_data = []
            
            for col in range(region['min_col'], region['max_col'] + 1):
                cell = sheet.cell(row, col)
                value = ""
                
                # Check if this cell is part of a merged range
                merged_value = self._get_merged_cell_value(sheet, row, col, merged_ranges)
                if merged_value is not None:
                    value = str(merged_value).strip()
                elif cell.value is not None:
                    value = str(cell.value).strip()
                
                # Apply formatting
                if cell.font and (cell.font.bold or cell.font.underline):
                    value = f"**{value}**" if value else ""
                
                row_data.append(value)
            
            data.append(row_data)
        
        return data
    
    def _get_merged_cell_value(self, sheet, row: int, col: int, merged_ranges) -> Optional[str]:
        """
        Get the value of a merged cell.
        
        Args:
            sheet: openpyxl worksheet object
            row (int): Row number
            col (int): Column number
            merged_ranges: List of merged cell ranges
            
        Returns:
            Optional[str]: Value of the merged cell or None
        """
        for merged_range in merged_ranges:
            if (row >= merged_range.min_row and row <= merged_range.max_row and
                col >= merged_range.min_col and col <= merged_range.max_col):
                # Return the value from the top-left cell of the merged range
                top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                return top_left_cell.value
        return None
    
    def _detect_multi_level_headers(self, data: List[List[str]]) -> Dict:
        """
        Detect and handle multi-level headers in table data.
        
        Args:
            data (List[List[str]]): Raw table data
            
        Returns:
            Dict: Processed data with header information
        """
        if not data:
            return {'headers': [], 'data_rows': [], 'header_levels': 0}
        
        max_cols = max(len(row) for row in data)
        header_levels = 0
        headers = []
        data_start_row = 0
        
        # Analyze first few rows to detect headers
        for i, row in enumerate(data[:5]):  # Check up to 5 rows for headers
            padded_row = row + [""] * (max_cols - len(row))
            
            # Check if this row looks like a header
            is_header = self._is_header_row(padded_row, i)
            
            if is_header:
                headers.append(padded_row)
                header_levels += 1
                data_start_row = i + 1
            else:
                break
        
        # If no headers detected, treat first row as header
        if header_levels == 0 and data:
            headers = [data[0] + [""] * (max_cols - len(data[0]))]
            header_levels = 1
            data_start_row = 1
        
        # Process remaining rows as data
        data_rows = []
        for row in data[data_start_row:]:
            padded_row = row + [""] * (max_cols - len(row))
            data_rows.append(padded_row)
        
        return {
            'headers': headers,
            'data_rows': data_rows,
            'header_levels': header_levels,
            'max_cols': max_cols
        }
    
    def _is_header_row(self, row: List[str], row_index: int) -> bool:
        """
        Determine if a row is likely a header row.
        
        Args:
            row (List[str]): Row data
            row_index (int): Index of the row
            
        Returns:
            bool: True if likely a header row
        """
        if row_index == 0:
            return True  # First row is usually a header
        
        # Count cells with bold formatting
        bold_count = sum(1 for cell in row if "**" in cell)
        total_non_empty = sum(1 for cell in row if cell.strip())
        
        # Header criteria
        if total_non_empty == 0:
            return False
        
        # If more than 50% of non-empty cells are bold, it's likely a header
        if bold_count > 0 and bold_count >= total_non_empty * 0.5:
            return True
        
        # Check for typical header patterns
        header_patterns = [
            r'^(total|sum|average|mean)$',
            r'^\w+\s+(total|sum|average)$',
            r'^(q[1-4]|quarter\s*[1-4])$',  # Quarterly headers
            r'^\d{4}$',  # Year headers
            r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)',  # Month headers
        ]
        
        pattern_matches = 0
        for cell in row:
            if cell.strip():
                cell_lower = cell.replace("**", "").lower().strip()
                for pattern in header_patterns:
                    if re.search(pattern, cell_lower):
                        pattern_matches += 1
                        break
        
        return pattern_matches >= max(1, total_non_empty * 0.3)
    
    def _create_markdown_with_multi_headers(self, processed_data: Dict) -> str:
        """
        Create markdown table with support for multi-level headers.
        
        Args:
            processed_data (Dict): Processed table data with header info
            
        Returns:
            str: Markdown formatted table
        """
        if not processed_data['headers']:
            return ""
        
        markdown_lines = []
        max_cols = processed_data['max_cols']
        
        # Handle multi-level headers
        if processed_data['header_levels'] > 1:
            # For multi-level headers, combine them into a single header row
            combined_headers = []
            
            for col_idx in range(max_cols):
                header_parts = []
                for header_level in processed_data['headers']:
                    if col_idx < len(header_level):
                        part = header_level[col_idx].replace("**", "").strip()
                        if part:
                            header_parts.append(part)
                
                # Combine header parts
                if header_parts:
                    combined_header = " - ".join(header_parts)
                else:
                    combined_header = f"Column {col_idx + 1}"
                
                combined_headers.append(combined_header)
            
            headers = combined_headers
        else:
            # Single level headers
            headers = [cell.replace("**", "").strip() or f"Column {i+1}" 
                      for i, cell in enumerate(processed_data['headers'][0])]
        
        # Create markdown table
        markdown_lines.append("| " + " | ".join(headers) + " |")
        markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        
        # Add data rows
        for row in processed_data['data_rows']:
            clean_row = [cell.replace("**", "").strip() for cell in row]
            markdown_lines.append("| " + " | ".join(clean_row) + " |")
        
        return "\n".join(markdown_lines)
    
def convert_excel_to_markdown(excel_file_path: str, 
                            sheet_name: Optional[str] = None,
                            output_file_path: Optional[str] = None,
                            specific_range: Optional[str] = None) -> str:
    """
    Convenience function to convert Excel to Markdown.
    
    Args:
        excel_file_path (str): Path to Excel file
        sheet_name (str, optional): Sheet name to convert
        output_file_path (str, optional): Path to save markdown file
        specific_range (str, optional): Specific range to convert (e.g., "A1:D10")
        
    Returns:
        str: Markdown content
    """
    converter = ExcelToMarkdownConverter(excel_file_path)
    
    try:
        if specific_range:
            markdown_content = converter.convert_specific_range(specific_range, sheet_name)
        else:
            markdown_content = converter.convert_sheet_to_markdown(sheet_name)
        
        if output_file_path:
            with open(output_file_path, 'w', encoding='utf-8') as f:
                f.write(markdown_content)
        
        return markdown_content
    
    finally:
        converter.close()


if __name__ == "__main__":
    # Example usage
    import os
    
    # Get the project root directory
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    
    # Example Excel file path
    excel_file = os.path.join(project_root, "sample_excel_data.xlsx")
    
    if os.path.exists(excel_file):
        print("Converting Excel file to Markdown...")
        print("=" * 50)
        
        # Convert entire sheet
        markdown_content = convert_excel_to_markdown(excel_file)
        print(markdown_content)
        
        # Save to file
        output_file = os.path.join(project_root, "converted_tables.md")
        convert_excel_to_markdown(excel_file, output_file_path=output_file)
        print(f"\nMarkdown saved to: {output_file}")
        
        # Example: Convert specific range
        # markdown_range = convert_excel_to_markdown(excel_file, specific_range="A1:C5")
        # print("\nSpecific range conversion:")
        # print(markdown_range)
    else:
        print(f"Sample Excel file not found at: {excel_file}")
        print("Please ensure you have a sample Excel file to test with.")
