import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

def convert_excel_to_html(excel_path: str, html_path: str):
    """
    Converts the first sheet of an Excel file to an HTML file,
    preserving styles like fonts, colors, borders, and alignment.

    Args:
        excel_path (str): The path to the input Excel file.
        html_path (str): The path to save the output HTML file.
    """
    try:
        # Load the workbook and select the first sheet
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # Start building the HTML content
        html_content = """
<!DOCTYPE html>
<html>
<head>
<style>
    body { font-family: Arial, sans-serif; }
    table { border-collapse: collapse; }
    th, td { border: 1px solid #ccc; padding: 8px; }
</style>
</head>
<body>
"""
        html_content += "<table>\n"

        # Set column widths
        for i in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(i)
            width = sheet.column_dimensions[col_letter].width
            if width:
                # Approximate conversion from Excel width to pixels
                html_content += f'<col style="width: {width * 9}px;">\n'

        # Iterate over rows and cells to build the table
        for row in sheet.iter_rows():
            html_content += "  <tr>\n"            
            for cell in row:
                # Start building inline CSS for the cell
                styles = []

                # Font styles
                if cell.font:
                    font = cell.font
                    if font.bold:
                        styles.append("font-weight: bold;")
                    if font.italic:
                        styles.append("font-style: italic;")
                    if font.size:
                        styles.append(f"font-size: {font.size}pt;")
                    if font.color and font.color.rgb:
                        try:
                            # Handle different types of color objects in openpyxl
                            color_hex = None
                            rgb_obj = font.color.rgb
                            
                            # Try different ways to extract hex color
                            if hasattr(rgb_obj, 'hex'):
                                color_hex = rgb_obj.hex
                            elif hasattr(rgb_obj, 'value'):
                                color_hex = rgb_obj.value
                            elif hasattr(rgb_obj, '__str__'):
                                color_str = str(rgb_obj)
                                # Only use if it looks like a hex color
                                if len(color_str) in [6, 8] and all(c in '0123456789ABCDEFabcdef' for c in color_str):
                                    color_hex = color_str
                            
                            if color_hex:
                                # Remove alpha channel if present (first 2 chars for ARGB)
                                if len(color_hex) == 8:
                                    color_hex = color_hex[2:]
                                styles.append(f"color: #{color_hex};")
                        except:
                            pass  # Skip color if extraction fails

                # Fill/Background color
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    try:
                        fill_hex = None
                        rgb_obj = cell.fill.fgColor.rgb
                        
                        # Try different ways to extract hex color
                        if hasattr(rgb_obj, 'hex'):
                            fill_hex = rgb_obj.hex
                        elif hasattr(rgb_obj, 'value'):
                            fill_hex = rgb_obj.value
                        elif hasattr(rgb_obj, '__str__'):
                            color_str = str(rgb_obj)
                            # Only use if it looks like a hex color
                            if len(color_str) in [6, 8] and all(c in '0123456789ABCDEFabcdef' for c in color_str):
                                fill_hex = color_str
                        
                        if fill_hex:
                            # Remove alpha channel if present
                            if len(fill_hex) == 8:
                                fill_hex = fill_hex[2:]
                            styles.append(f"background-color: #{fill_hex};")
                    except:
                        pass  # Skip if color extraction fails

                # Alignment
                if cell.alignment:
                    align = cell.alignment
                    if align.horizontal:
                        styles.append(f"text-align: {align.horizontal};")
                    if align.vertical:
                        styles.append(f"vertical-align: {align.vertical};")                # Border styles
                if cell.border:
                    border = cell.border
                    for side_name in ['left', 'right', 'top', 'bottom']:
                        side = getattr(border, side_name)
                        if side and side.style:
                            color_hex = '000000'  # default to black
                            if side.color and side.color.rgb:
                                try:
                                    rgb_str = str(side.color.rgb)
                                    color_hex = rgb_str[2:] if len(rgb_str) == 8 else rgb_str
                                except:
                                    color_hex = '000000'
                            styles.append(f"border-{side_name}: 1px solid #{color_hex};")

                style_attr = f'style="{" ".join(styles)}"' if styles else ""
                
                # Get cell value
                cell_value = cell.value if cell.value is not None else ""

                html_content += f'    <td {style_attr}>{cell_value}</td>\n'
            html_content += "  </tr>\n"

        html_content += "</table>\n"
        html_content += "</body>\n</html>"

        # Write the HTML content to the specified file
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
            
        print(f"Successfully converted '{excel_path}' to '{html_path}'")

    except FileNotFoundError:
        print(f"Error: The file '{excel_path}' was not found.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == '__main__':
    # Example Usage:
    # Create a dummy Excel file for testing purposes.
    
    # Define paths
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    sample_excel_path = os.path.join(project_root, "sample_excel_data.xlsx")
    output_html_path = os.path.join(project_root, "sample_excel_output.html")

    # Check if sample exists, otherwise create it
    if not os.path.exists(sample_excel_path):
        print("Creating a sample styled Excel file for demonstration...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Styled Report"

        # Header
        ws['A1'] = "Financial Report"
        ws.merge_cells('A1:C1')
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill("solid", fgColor="4F81BD")
        ws['A1'].alignment = Alignment(horizontal='center')

        # Data Headers
        headers = ["Metric", "Value", "Change"]
        ws.append(headers)
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DCE6F1")

        # Data
        data = [
            ("Revenue", 150000, "5%"),
            ("Expenses", 80000, "2%"),
            ("Profit", 70000, "10%")
        ]
        for row_data in data:
            ws.append(row_data)

        # Apply styles to data
        for row in ws.iter_rows(min_row=3, max_row=5, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal='right')

        # Add a border around the profit cell
        ws['B5'].border = Border(left=Side(style='thin', color='FF0000'), 
                                 right=Side(style='thin', color='FF0000'), 
                                 top=Side(style='thin', color='FF0000'), 
                                 bottom=Side(style='thin', color='FF0000'))
        
        wb.save(sample_excel_path)
        print(f"Sample file created at '{sample_excel_path}'")

    # Run the conversion
    print("\n--- Running Excel to HTML Conversion ---")
    convert_excel_to_html(sample_excel_path, output_html_path)
